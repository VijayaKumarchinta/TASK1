"""
Utility functions for the Netflix Data Cleaning Pipeline.
Handles logging, reporting, and helper utilities.
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd


class NumpyEncoder(json.JSONEncoder):
    """Custom JSON encoder that handles numpy types."""
    def default(self, obj):
        try:
            import numpy as np
            if isinstance(obj, (np.integer,)):
                return int(obj)
            if isinstance(obj, (np.floating,)):
                return float(obj)
            if isinstance(obj, (np.ndarray,)):
                return obj.tolist()
            if isinstance(obj, (np.bool_,)):
                return bool(obj)
        except ImportError:
            pass
        return super().default(obj)


class CleaningLogger:
    """Logs all cleaning operations performed on the dataset."""

    def __init__(self):
        self.log: List[Dict[str, Any]] = []
        self.start_time = datetime.now()

    def log_operation(
        self,
        operation: str,
        description: str,
        affected_rows: Optional[int] = None,
        details: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Log a cleaning operation with metadata."""
        entry = {
            "timestamp": datetime.now().isoformat(),
            "operation": operation,
            "description": description,
            "affected_rows": affected_rows,
            "details": details or {},
        }
        self.log.append(entry)
        print(f"  [{operation}] {description}", end="")
        if affected_rows is not None:
            print(f" ({affected_rows} rows affected)")
        else:
            print()

    def generate_report(self) -> str:
        """Generate a formatted summary report of all operations."""
        elapsed = datetime.now() - self.start_time
        lines = [
            "=" * 70,
            "  DATA CLEANING SUMMARY REPORT",
            "=" * 70,
            f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"  Total operations: {len(self.log)}",
            f"  Duration: {elapsed.total_seconds():.2f} seconds",
            "-" * 70,
            "",
        ]

        for i, entry in enumerate(self.log, 1):
            lines.append(f"  Step {i}: {entry['operation']}")
            lines.append(f"    Description: {entry['description']}")
            if entry["affected_rows"] is not None:
                lines.append(f"    Rows affected: {entry['affected_rows']:,}")
            if entry["details"]:
                for key, value in entry["details"].items():
                    lines.append(f"    {key}: {value}")
            lines.append("")

        lines.extend(["-" * 70, "  END OF REPORT", "=" * 70])
        return "\n".join(lines)

    def save_report(self, filepath: str) -> None:
        """Save the report to a text file."""
        report = self.generate_report()
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"\n[DOC] Report saved to: {filepath}")

    def save_json_log(self, filepath: str) -> None:
        """Save the detailed JSON log."""
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "start_time": self.start_time.isoformat(),
                    "end_time": datetime.now().isoformat(),
                    "operations": self.log,
                },
                f,
                indent=2,
                cls=NumpyEncoder,
            )
        print(f"[LIST] JSON log saved to: {filepath}")


def get_hashable_cols(df) -> list:
    """Return columns that contain only hashable types (not lists/dicts)."""
    return [c for c in df.columns
            if df[c].apply(lambda x: isinstance(x, (list, dict))).sum() == 0]


def export_to_excel(df, filepath: str, sheet_name: str = "Netflix_Cleaned") -> str:
    """
    Export a DataFrame to a nicely formatted Excel file.

    Features:
      - Auto-adjusted column widths
      - Bold header row with Netflix-red background
      - White text on headers
      - Alternating row colors for readability
      - Freeze panes on header row
      - Auto-filter enabled on all columns
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARN] openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    ensure_dir(os.path.dirname(filepath))

    # Write to Excel with openpyxl engine for full formatting control
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # --- Styling constants ---
        header_fill = PatternFill(start_color="E50914", end_color="E50914", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )
        even_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # --- Style header row ---
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # --- Style data rows (alternating colors + borders) ---
        for row_idx in range(2, len(df) + 2):
            fill = even_fill if row_idx % 2 == 0 else odd_fill
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # --- Auto-fit column widths (capped at 50) ---
        for col_idx, col_name in enumerate(df.columns, 1):
            # Get max length in column (header + data)
            max_len = len(str(col_name))
            for row_idx in range(2, min(len(df) + 2, 100)):  # Sample first 100 rows for speed
                cell_val = str(worksheet.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(cell_val))
            adjusted_width = min(max_len + 2, 50)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # --- Freeze header row & add auto-filter ---
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    file_size_kb = os.path.getsize(filepath) / 1e3
    print(f"  [EXCEL] Saved styled Excel -> {filepath} ({file_size_kb:.1f} KB)")
    return filepath


def ensure_dir(path: str) -> None:
    """Create directory if it doesn't exist."""
    Path(path).mkdir(parents=True, exist_ok=True)


def describe_dataframe_basics(df_summary: Any) -> str:
    """Create a text summary of DataFrame info for the report."""
    lines = [
        "DATASET OVERVIEW",
        "-" * 50,
        f"  Rows: {df_summary.get('rows', 'N/A'):,}",
        f"  Columns: {df_summary.get('cols', 'N/A')}",
        f"  Memory usage: {df_summary.get('memory_mb', 0):.2f} MB",
        "",
        "COLUMN DATA TYPES:",
    ]
    for col, dtype in df_summary.get("dtypes", {}).items():
        lines.append(f"  ? {col}: {dtype}")
    lines.append("")
    lines.append("MISSING VALUES (BEFORE CLEANING):")
    for col, count in df_summary.get("nulls_before", {}).items():
        if count > 0:
            pct = (count / df_summary["rows"]) * 100
            lines.append(f"  ? {col}: {count:,} ({pct:.1f}%)")
    lines.append("")
    lines.append("DUPLICATES:")
    lines.append(f"  ? Duplicate rows: {df_summary.get('duplicates', 0):,}")
    return "\n".join(lines)
